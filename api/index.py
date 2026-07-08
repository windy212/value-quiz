import os, json, time, io
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)

DATA_FILE = '/tmp/responses.json'
ADMIN_PASSWORD = os.environ.get('ADMIN_PWD', 'admin123')

def _load():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r') as f:
                return json.load(f)
        except:
            return []
    return []

def _save(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, ensure_ascii=False)

@app.route('/submit', methods=['POST'])
def submit():
    data = request.get_json()
    if not data or 'answers' not in data:
        return jsonify({'error': 'bad'}), 400
    answers = data['answers']
    name = data.get('name', '\u533f\u540d')
    groups = []
    for g in range(13):
        start, end = g * 8, min(g * 8 + 8, 99)
        score = sum(int(answers.get(str(i + 1), 0)) for i in range(start, end))
        groups.append(score)
    record = {
        'id': int(time.time() * 1000),
        'name': name,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'groups': groups,
        'total_score': sum(groups)
    }
    all_data = _load()
    all_data.append(record)
    _save(all_data)
    return jsonify({'success': True})

@app.route('/admin', methods=['POST'])
def admin():
    password = request.form.get('password') or request.args.get('password')
    if password != ADMIN_PASSWORD:
        return jsonify({'error': 'wrong password'}), 401
    return jsonify({'responses': _load(), 'count': _load().__len__()})

@app.route('/export')
def export():
    password = request.args.get('password')
    if password != ADMIN_PASSWORD:
        return 'Unauthorized', 401
    all_data = _load()
    if not all_data:
        return 'no data', 200
    wb = Workbook()
    ws = wb.active
    ws.title = 'summary'
    hfont = Font(bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    b = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ['\u5e8f\u53f7', '\u59d3\u540d', '\u63d0\u4ea4\u65f6\u95f4'] + [f'\u7b2c{g}\u7ec4' for g in range(1, 14)] + ['\u603b\u5206']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, b
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    for col in range(4, 18):
        ws.column_dimensions[chr(64 + col)].width = 8
    ws.column_dimensions['Q'].width = 8
    for idx, resp in enumerate(all_data, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = b
        ws.cell(row=row, column=2, value=resp.get('name', 'N/A')).border = b
        ws.cell(row=row, column=3, value=resp['timestamp']).border = b
        for gi, s in enumerate(resp.get('groups', [])):
            c = ws.cell(row=row, column=4 + gi, value=s)
            c.border, c.alignment = b, Alignment(horizontal='center')
        tc = ws.cell(row=row, column=17, value=resp.get('total_score', 0))
        tc.border, tc.font, tc.alignment = b, Font(bold=True, color='D93025'), Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='export.xlsx')
